# -*- coding: utf-8 -*-

import shutil
import sqlite3
import re
import os
import sys
import yaml
from openpyxl import load_workbook
import xlsxwriter
import itertools
import time
import json
import demjson

def Sheet2List(sheet):
	listResult = []
	for i in range(1,sheet.max_row + 1):
		lineData = []
		for j in range(1,sheet.max_column +1):
			cell = sheet.cell(row = i, column = j)
			lineData.append(cell.value)
		listResult.append(lineData)
	return listResult

def readJson():
	print "start read json"
	data = demjson.decode(open("CrossDataaa.json", 'r').read())
	print "finish read json"
	# print data
	return data

def saveJson(data):
	with open('CrossDataaa.json', 'w') as result_file:
		json.dump(data, result_file)

def getSourceList(sourceName,sheetName):
    ws = load_workbook(sourceName)
    sheets = ws.sheetnames
    bookSheet = ws[sheetName]

    rows = bookSheet.rows
    columns = bookSheet.columns

    return Sheet2List(bookSheet)


def readXlsx():
	wordSourceList = getSourceList('new1203.xlsx',"word")
	wordSourceMap = {}
	for i in range(1,len(wordSourceList)):
		wordSource = wordSourceList[i]
		if wordSource[0] is not None:
			wordSourceMap[wordSource[1]] = wordSource[0]

	return wordSourceMap

def getLetterList(letterList,wordSourceMap):
	words = []
	for i in range(3,len(letterList)+1):
		wordList = list(itertools.permutations(letterList,i))
		for x in wordList:
			word = ""
			for y in x:
				word = word + y
			words.append(word)
    #去重操作
	wordsList = []
	for x in words:
		if x not in wordsList:
			#判断是否是 单词
			if wordSourceMap.get(x):
				print x
				wordsList.append(x)

	return wordsList

def write_cross_data(filename, resultArr):
	fd = open(filename, "wb")
	fd.write("module.exports = {\n")
	fd.write("\"plots\" : {\n")

	for index in range(len(resultArr)):
		data = resultArr[str(index+1)]
		fd.write("\"{}\" : {},\n".format(str(index+1), json.dumps(data)))
	fd.write("}\n}\n")
	fd.close()


def addExtra():
	originCrossData = readJson()
	print "read xlsx"
	wordSourceMap = readXlsx()
	print "finish read xlsx"
	for x in originCrossData:
		words = getLetterList(originCrossData[x]["letters"],wordSourceMap)
		coordinates = []
		for word in originCrossData[x]["coordinates"]:
			coordinates.append(word[0])
		print '------------------------'
		#判断是否 是正常单词
		for word in words:
			if word not in coordinates:
				if originCrossData[x].get("extra"):
					if word not in originCrossData[x]["extra"]:
						originCrossData[x]["extra"].append(word)
				else:
					originCrossData[x]["extra"] = []
					originCrossData[x]["extra"].append(word)
	write_cross_data("newWordCrossData.ts", originCrossData)

# addExtra()

# resultArr = readJson()
# write_cross_data("newWordCrossData.ts", resultArr)


































