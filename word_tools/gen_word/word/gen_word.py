# -*- coding: utf-8 -*-

import shutil
import sqlite3
import re
import os
import sys
import yaml
from openpyxl import load_workbook
import xlsxwriter
import itertools
import time
import json

CIPIN_A = 10000
CIPIN_B = 10000

MIN_WORDLETTER_NUM = 3

OUT_NO_SKIP_WORD_FILENAME = "gen_word/word/nohave_2_spip1_reset_spkip2_new.xlsx"

ORIGINAL_WORD_FILENAME = "gen_word/word/cipinbiao.xlsx"
#必须去掉的屏蔽词
PINGBICI_MUST_REMOVE_FILENAME = "gen_word/word/pingbici_1.txt"
#重置排名最大的屏蔽词
PINGBICI_RESET_RANK_FILENAME = "gen_word/word/pingbici_2.txt"

OUT_CIPIN_WORD_FILENAME = "gen_word/word/word_10000w.xlsx"

# 把一个表格中的数据全部导出到一个列表
def Sheet2List(sheet):
    listResult = []
    for i in range(1,sheet.max_row + 1):
        lineData = []
        for j in range(1,sheet.max_column +1):
            cell = sheet.cell(row = i, column = j)
            lineData.append(cell.value)
        listResult.append(lineData)

    return listResult

def getSourceList(sourceName,sheetName):
    ws = load_workbook(sourceName)
    sheets = ws.sheetnames
    bookSheet = ws[sheetName]

    rows = bookSheet.rows
    columns = bookSheet.columns

    return Sheet2List(bookSheet)


def getRemovePingbici():
    skipListContent = open(PINGBICI_MUST_REMOVE_FILENAME, "r").read()
    skipList = skipListContent.split()
    return skipList

def getResetRankPingbici():
    skipListContent = open(PINGBICI_RESET_RANK_FILENAME, "r").read()
    skipList = skipListContent.split()
    return skipList

def genNewWord():
    fileName = OUT_NO_SKIP_WORD_FILENAME
    if(os.path.exists(fileName)):
        os.remove(fileName)
    newXlsx = xlsxwriter.Workbook(fileName)
    skipList = getRemovePingbici()
    resetRankList = getResetRankPingbici()
    sht1 = newXlsx.add_worksheet("word")
    sht1.write(0,1,"Word")
    sht1.write(0,0,"Rank")
    wordSourceList = getSourceList(ORIGINAL_WORD_FILENAME,"word")
    index = 1
    for x in range(1,len(wordSourceList)):
        wordSource = wordSourceList[x]
        word = str(wordSource[1])
        if word.isalpha() and word.lower() not in skipList:
            if word is not None and len(word) > 1:
                rank = wordSource[0]
                if word.lower() in resetRankList:
                    rank = 99999
                sht1.write(index,1,str(wordSource[1]).upper())
                sht1.write(index,0,rank)
                index = index + 1         
    newXlsx.close()    
    print "gen_newword success!!!"

def getWordKey(word):
    wordlist = list(str(word))
    word = "".join((lambda x:(x.sort(),x)[1])(wordlist))
    wordkey = ""
    for i in range(0,len(wordlist)):
        if i >= len(wordlist)-1:
            wordkey = wordkey + wordlist[i]
        else:
            wordkey = wordkey + wordlist[i] + "|"

    return wordkey

def getLetterList(word):
    words = []
    letterList = list(word)
    for i in range(MIN_WORDLETTER_NUM,len(letterList)+1):
        wordList = list(itertools.permutations(letterList,i))
        for x in wordList:
            word = ""
            for y in x:
                word = word + y
            words.append(word)

    #去重操作
    wordList = []
    for x in words:
        if x not in wordList:
            wordList.append(x)

    return wordList

def changeListToStr(wordlist):
    wordStr = ""
    for i in range(0,len(wordlist)):
        if i >= len(wordlist)-1:
            wordStr = wordStr + wordlist[i]
        else:
            wordStr = wordStr + wordlist[i] + "|"

    return wordStr

def logWordXlsx(resultList):
    fileName = OUT_CIPIN_WORD_FILENAME
    if(os.path.exists(fileName)):
        os.remove(fileName)
    workbook = xlsxwriter.Workbook(fileName)
    worksheet = workbook.add_worksheet("word")
    worksheet.write(0,3,"extraWord")
    worksheet.write(0,2,"keyWord")
    worksheet.write(0,1,"wordNum")
    worksheet.write(0,0,"setLetter")
    index = 1;
    for key in resultList:
        worksheet.write(index,3,resultList[key]["extraWord"])
        worksheet.write(index,2,resultList[key]["keyWord"])
        worksheet.write(index,1,(len(key)+1)/2)
        worksheet.write(index,0,key)
        index = index + 1

    workbook.close()
    print("log wordXlsx  success!!!!!!")

def checkResultWord(word,resultList):
    wordLen = len(word)
    listLen = len(resultList)
    if listLen <= 1:
        return False

    #2个字母组合需要至少2个答案词，3个和4个字母组合需要至少3个答案词，5个和6个字母组合至少需要4个答案词
    if wordLen == 2:
        if listLen >= 2:
            return True
    elif wordLen == 3 or wordLen == 4:
        if listLen >= 3:
            return True
    elif wordLen >= 5:
        if listLen >= 4:
            return True

    return False


def genCipinWord():
    wordSourceList = getSourceList(OUT_NO_SKIP_WORD_FILENAME,"word")
    wordSourceMap = {}
    for i in range(1,len(wordSourceList)):
        wordSource = wordSourceList[i]
        if wordSource[0] is not None:
            wordSourceMap[wordSource[1]] = wordSource[0]

    resultList = {}
    num = 0
    for i in range(1,len(wordSourceList)):
        wordSource = wordSourceList[i]
        if not wordSource[1] is None and not wordSource[0] is None and int(wordSource[0]) <= CIPIN_A:
            wordkey = getWordKey(wordSource[1])
            print wordkey

            if not resultList.get(wordkey):
                wordList = getLetterList(wordSource[1])
                rightwordlist = []
                extraWordList = []
                for word in wordList:
                    if wordSourceMap.get(word):
                        if wordSourceMap[word] >= CIPIN_B:
                            extraWordList.append(word)
                        else:
                            rightwordlist.append(word)
                if checkResultWord(wordSource[1],rightwordlist): 
                    num = num + 1
                    resultList[wordkey] = {"keyWord":changeListToStr(rightwordlist),"extraWord":changeListToStr(extraWordList)}
    print("solv success!!!!",num)
    logWordXlsx(resultList)

def getJsonSourceData(fileName):
    fileName = "gen_word/CompetingGoodsAnalysis/wordscape/wordgroup_all/" + fileName
    print fileName
    if(os.path.exists(fileName)):
        f = open(fileName, 'r').read() 
        data = json.loads(f) 
        return data

def getMaxLenWordNumAndMeanLetters(levelData):
    wordList = levelData["e"]
    maxLen = 0
    allLen = 0

    lettersNum = len(levelData["b"])
    for words in wordList:
        word = words.split(",")[3]
        allLen += len(word)
        if lettersNum == len(word):
            maxLen += 1
           
    meanLen = round(allLen*1.0 / len(wordList),2)

    return maxLen,meanLen

def logAnalyticWordScape(dataList):
    fileName = "gen_word/CompetingGoodsAnalysis/wordscape/wordscape_all_analytic.xlsx"
    if(os.path.exists(fileName)):
        os.remove(fileName)
    workbook = xlsxwriter.Workbook(fileName)
    worksheet = workbook.add_worksheet("word")
    worksheet.write(0,5,"avg_letters")
    worksheet.write(0,4,"longest_words")
    worksheet.write(0,3,"words")
    worksheet.write(0,2,"letters")
    worksheet.write(0,1,"level")
    worksheet.write(0,0,"json")
    index = 1;
    for data in dataList:
        worksheet.write(index,5,data["avg_letters"])
        worksheet.write(index,4,data["longest_words"])
        worksheet.write(index,3,data["words"])
        worksheet.write(index,2,data["letters"])
        worksheet.write(index,1,data["level"])
        worksheet.write(index,0,data["json"])
        index = index + 1

    workbook.close()
    print("log wordXlsx  success!!!!!!")   

def analyticWordScape():
    parents = os.listdir("gen_word/CompetingGoodsAnalysis/wordscape/wordgroup_all")
    dataList = []
    for child in parents:
        print child
        groupData = getJsonSourceData(child)
        fileName = child.split(".")[0]
        keys = groupData.keys()
        keys.sort()
        for level in keys:
            if level:
                childDict = {}
                levelData = groupData[level]
                childDict["json"] = child.split(".")[0]
                childDict["level"] = level
                childDict["letters"] = len(levelData["b"])
                childDict["words"] = len(levelData["e"])
                maxLenNum,meanLen = getMaxLenWordNumAndMeanLetters(levelData)
                childDict["longest_words"] = maxLenNum
                childDict["avg_letters"] = meanLen
                dataList.append(childDict)

    logAnalyticWordScape(dataList);

#去掉屏蔽词和两个单词的原始数据文件
# genNewWord()

#生成词频表 并且根据次品分类
# genCipinWord()

#分析竞品工具
# analyticWordScape()











