# -*- coding: utf-8 -*-
import shutil
import sqlite3
import re
import os
import sys
import yaml
from openpyxl import load_workbook
import xlsxwriter
import itertools
import time
import json

def Sheet2List(sheet):
    listResult = []
    for i in range(1,sheet.max_row + 1):
        lineData = []
        for j in range(1,sheet.max_column +1):
            cell = sheet.cell(row = i, column = j)
            lineData.append(cell.value)
        listResult.append(lineData)

    return listResult

#txt 文件读取 返回 list
def readTxtFile_List(filePath):
	fileData = open(filePath, "r").read()
	skipList = fileData.split()
	return skipList

#xlsx 文件读取 返回文件数据

def readXlsxFile_data(filePath,sheetName):
	ws = load_workbook(filePath)
	if sheetName == 0:
		sheets = ws.sheetnames
		sheetName = sheets[0]
    
	bookSheet = ws[sheetName]
	return bookSheet

#xlsx 文件读取 返回list
def readXlsxFile_List(filePath,sheetName = 0):
	bookSheet = readXlsxFile_data(filePath,sheetName)

	listResult = []
	for i in range(1,bookSheet.max_row + 1):
		lineData = []
		for j in range(1,bookSheet.max_column +1):
			cell = bookSheet.cell(row = i, column = j)
			lineData.append(cell.value)
		listResult.append(lineData)

	return listResult

#xlsx 文件读取 返回list
def readXlsxFile_Map(filePath,sheetName = 0):
	bookSheet = readXlsxFile_data(filePath,sheetName)

	mapResult = {}

	for i in range(1,bookSheet.max_column + 1):
		cell = bookSheet.cell(1, column = i)
		mapResult[cell.value] = []

	for i in range(1,bookSheet.max_column + 1):
		key = bookSheet.cell(1, column = i).value
		for j in range(2,bookSheet.max_row +1):
			cell = bookSheet.cell(row = j, column = i)
			# lineData.append(cell.value)
			mapResult[key].append(cell.value)

	return mapResult


def readJsonFile(fileName):
	if(os.path.exists(fileName)):
		print(fileName)
		jsonContent = open(fileName, 'r').read()
		return json.loads(jsonContent)
	return None
#json 文件读取
# def readJsonFile(fileName):
# 	if(os.path.exists(fileName)):

# 		with open(fileName,'r') as load_f:
# 			return json.load(load_f)

# 	return

#json 文件写入
def saveJsonFile(fileName,data):
	json_str = json.dumps(data, indent=4)
	with open(fileName,"w") as json_file:
		json_file.write(json_str)	


def parse_result_to_ts_data(fileName,data):
	fd = open(filename, "wb")
	fd.write("const CrossData = {\n")
	fd.write( json.dumps(data))
	fd.write("}\nexport {CrossData}")
	fd.close()

def write_cross_data_for_map(filename, resultData):
	fd = open(filename, "wb")
	fd.write("module.exports = {\n")
	for key in resultData:
		data = resultData[key]
		fd.write("\"{}\" : {},\n".format(key, json.dumps(data)))
	fd.write("\n}\n")
	fd.close()

def write_cross_data_for_list(filename, resultData):
	fd = open(filename, "wb")
	fd.write("const CrossData2 = {\n")
	index = 1
	for key in resultData:
		data = key #resultData[key]
		fd.write("\"{}\" : {},\n".format(str(index), json.dumps(data)))
		index = index + 1
	fd.write("\n}\n export { CrossData2 };")
	fd.close()
		




